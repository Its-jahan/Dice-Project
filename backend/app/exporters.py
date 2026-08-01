"""Export a holder result set as CSV, XLSX, JSON or a standalone HTML page."""

from __future__ import annotations

import csv
import html
import io
import json
from datetime import date, datetime, timezone

from .models import ExportFormat, HolderMode, HoldersResponse

SNAPSHOT_HEADERS = ["wallet_address", "token_address", "snapshot_date", "balance"]
SUMMARY_HEADERS = [
    "wallet_address",
    "first_date",
    "last_date",
    "days_held",
    "min_balance",
    "max_balance",
    "avg_balance",
]

MEDIA_TYPES = {
    ExportFormat.csv: "text/csv; charset=utf-8",
    ExportFormat.xlsx: (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ),
    ExportFormat.json: "application/json",
    ExportFormat.html: "text/html; charset=utf-8",
}


def filename_for(
    result: HoldersResponse, fmt: ExportFormat, dataset: str = "auto"
) -> str:
    req = result.request
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    token = req.token_address[:10]
    table = "summary" if _is_summary_first(result, dataset) else "snapshots"
    return (
        f"dice_{req.chain.value}_{token}_{req.start_date}_{req.effective_end_date}"
        f"_{req.holder_mode.value}_{table}_{stamp}.{fmt.value}"
    )


def export(
    result: HoldersResponse, fmt: ExportFormat, dataset: str = "auto"
) -> bytes:
    if fmt is ExportFormat.csv:
        return _to_csv(result, dataset)
    if fmt is ExportFormat.json:
        return _to_json(result)  # JSON always carries both tables
    if fmt is ExportFormat.html:
        return _to_html(result, dataset)
    return _to_xlsx(result, dataset)


def _snapshot_rows(result: HoldersResponse) -> list[list[object]]:
    return [
        [s.wallet_address, s.token_address, s.snapshot_date.isoformat(), s.balance]
        for s in result.snapshots
    ]


def _summary_rows(result: HoldersResponse) -> list[list[object]]:
    return [
        [
            s.wallet_address,
            s.first_date.isoformat(),
            s.last_date.isoformat(),
            s.days_held,
            s.min_balance,
            s.max_balance,
            round(s.avg_balance, 8),
        ]
        for s in result.summary
    ]


#: Which table an export leads with. "auto" follows the holder mode; the
#: explicit values let the UI download whichever tab the user is looking at.
DATASETS = ("auto", "snapshots", "summary")


def _is_summary_first(result: HoldersResponse, dataset: str = "auto") -> bool:
    """Whether the summary table leads this export.

    "auto" follows the holder mode: any_time and continuous are wallet-level
    questions, daily is row-level. An explicit dataset overrides that, so the
    Download button can honour the tab the user is actually viewing.
    """
    if dataset == "summary":
        return True
    if dataset == "snapshots":
        return False
    return result.request.holder_mode is not HolderMode.daily


def _to_csv(result: HoldersResponse, dataset: str = "auto") -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    if _is_summary_first(result, dataset):
        writer.writerow(SUMMARY_HEADERS)
        writer.writerows(_summary_rows(result))
    else:
        writer.writerow(SNAPSHOT_HEADERS)
        writer.writerows(_snapshot_rows(result))
    return buffer.getvalue().encode("utf-8-sig")


def _to_json(result: HoldersResponse) -> bytes:
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "request": json.loads(result.request.model_dump_json()),
        "execution_id": result.execution_id,
        "row_count": result.row_count,
        "wallet_count": result.wallet_count,
        "truncated": result.truncated,
        "snapshots": [json.loads(s.model_dump_json()) for s in result.snapshots],
        "summary": [json.loads(s.model_dump_json()) for s in result.summary],
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


def _to_xlsx(result: HoldersResponse, dataset: str = "auto") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheets: list[tuple[str, list[str], list[list[object]]]] = [
        ("Snapshots", SNAPSHOT_HEADERS, _snapshot_rows(result)),
        ("Summary", SUMMARY_HEADERS, _summary_rows(result)),
    ]
    if _is_summary_first(result, dataset):
        sheets.reverse()

    first = True
    for title, headers, rows in sheets:
        sheet = workbook.active if first else workbook.create_sheet()
        sheet.title = title
        first = False
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 1)}"
        )
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(
                14, len(header) + 4
            )

    meta = workbook.create_sheet("Request")
    meta.append(["field", "value"])
    meta["A1"].font = Font(bold=True)
    meta["B1"].font = Font(bold=True)
    for key, value in json.loads(result.request.model_dump_json()).items():
        meta.append([key, str(value)])
    meta.append(["execution_id", result.execution_id or ""])
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 48

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _to_html(result: HoldersResponse, dataset: str = "auto") -> bytes:
    req = result.request
    summary_first = _is_summary_first(result, dataset)
    rows = _summary_rows(result) if summary_first else _snapshot_rows(result)
    headers = SUMMARY_HEADERS if summary_first else SNAPSHOT_HEADERS
    title = f"DICE — {req.chain.value} — {req.token_address}"

    body_rows = "\n".join(
        "<tr>" + "".join(f"<td>{html.escape(_cell(c))}</td>" for c in row) + "</tr>"
        for row in rows
    )
    head_cells = "".join(f"<th>{html.escape(h)}</th>" for h in headers)

    document = f"""<!doctype html>
<html lang="en">
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(title)}</title>
<style>
  :root {{ color-scheme: light dark; }}
  body {{ font: 15px/1.5 ui-sans-serif, system-ui, sans-serif; margin: 2rem; }}
  h1 {{ font-size: 1.25rem; margin-bottom: .25rem; }}
  .meta {{ color: #6b7280; font-size: .85rem; margin-bottom: 1rem; }}
  input {{ padding: .5rem .75rem; width: min(420px, 100%); margin-bottom: 1rem;
           border: 1px solid #9ca3af; border-radius: .5rem; font: inherit; }}
  .scroll {{ overflow-x: auto; }}
  table {{ border-collapse: collapse; width: 100%; font-variant-numeric: tabular-nums; }}
  th, td {{ border-bottom: 1px solid #d1d5db; padding: .4rem .6rem; text-align: left;
            white-space: nowrap; }}
  th {{ position: sticky; top: 0; background: Canvas; }}
</style>
<h1>{html.escape(title)}</h1>
<div class="meta">
  {html.escape(req.start_date.isoformat())} → {html.escape(req.end_date.isoformat())}
  (UTC) · mode: {html.escape(req.holder_mode.value)} ·
  min balance: {req.min_balance} · {result.wallet_count} wallets ·
  {result.row_count} snapshot rows{" · TRUNCATED" if result.truncated else ""}
</div>
<input id="q" type="search" placeholder="Filter by wallet address…"
       aria-label="Filter by wallet address">
<div class="scroll">
<table id="t">
  <thead><tr>{head_cells}</tr></thead>
  <tbody>
{body_rows}
  </tbody>
</table>
</div>
<script>
  const input = document.getElementById('q');
  const rows = Array.from(document.querySelectorAll('#t tbody tr'));
  input.addEventListener('input', () => {{
    const needle = input.value.trim().toLowerCase();
    for (const row of rows) {{
      row.hidden = needle !== '' && !row.cells[0].textContent.toLowerCase().includes(needle);
    }}
  }});
</script>
</html>
"""
    return document.encode("utf-8")


def _cell(value: object) -> str:
    if isinstance(value, float):
        return f"{value:,.8f}".rstrip("0").rstrip(".")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)
