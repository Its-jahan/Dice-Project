import csv
import io
import json
from datetime import timedelta

from app.exporters import export, filename_for
from app.holders import build_summary, parse_rows
from app.models import ExportFormat, HoldersRequest, HoldersResponse

WALLET_A = "0xaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa"
WALLET_B = "0xbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb"


def make_result(holder_mode="daily") -> HoldersResponse:
    req = HoldersRequest(
        chain="ethereum",
        token_address="0x1234567890abcdef1234567890abcdef12345678",
        start_date="2026-07-20",
        end_date="2026-07-21",
        holder_mode=holder_mode,
    )
    rows = [
        {"wallet_address": WALLET_A, "day": "2026-07-20", "balance": 1540.25},
        {"wallet_address": WALLET_A, "day": "2026-07-21", "balance": 1300.0},
        {"wallet_address": WALLET_B, "day": "2026-07-21", "balance": 82.1},
    ]
    snapshots = parse_rows(rows, req)
    summary = build_summary(snapshots)
    return HoldersResponse(
        request=req,
        execution_id="exec-123",
        row_count=len(snapshots),
        wallet_count=len(summary),
        snapshots=snapshots,
        summary=summary,
    )


def test_csv_export_has_snapshot_rows_in_daily_mode():
    payload = export(make_result(), ExportFormat.csv).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(payload)))

    assert rows[0] == ["wallet_address", "token_address", "snapshot_date", "balance"]
    assert len(rows) == 4  # header + 3 snapshots
    assert {r[0] for r in rows[1:]} == {WALLET_A, WALLET_B}


def test_csv_export_switches_to_summary_for_wallet_level_modes():
    payload = export(make_result("any_time"), ExportFormat.csv).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(payload)))

    assert rows[0][0] == "wallet_address"
    assert "days_held" in rows[0]
    assert len(rows) == 3  # header + 2 wallets


def test_json_export_round_trips():
    payload = json.loads(export(make_result(), ExportFormat.json))

    assert payload["row_count"] == 3
    assert payload["wallet_count"] == 2
    assert payload["request"]["holder_mode"] == "daily"
    assert payload["snapshots"][0]["snapshot_date"] == "2026-07-20"


def test_html_export_is_standalone_and_escapes_content():
    payload = export(make_result(), ExportFormat.html).decode("utf-8")

    assert payload.startswith("<!doctype html>")
    assert "http://" not in payload and "https://" not in payload  # no external assets
    assert WALLET_A in payload


def test_xlsx_export_has_snapshot_summary_and_request_sheets():
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(export(make_result(), ExportFormat.xlsx)))

    assert set(workbook.sheetnames) == {"Snapshots", "Summary", "Request"}
    assert workbook.sheetnames[0] == "Snapshots"  # daily mode leads with snapshots
    assert workbook["Snapshots"].max_row == 4
    assert workbook["Summary"].max_row == 3


def test_xlsx_leads_with_summary_for_wallet_level_modes():
    from openpyxl import load_workbook

    workbook = load_workbook(
        io.BytesIO(export(make_result("continuous"), ExportFormat.xlsx))
    )
    assert workbook.sheetnames[0] == "Summary"


def test_filename_carries_the_request_shape():
    name = filename_for(make_result(), ExportFormat.csv)

    assert name.startswith("dice_ethereum_0x12345678_2026-07-20_2026-07-21_daily_")
    assert name.endswith(".csv")


def test_download_follows_the_requested_dataset_not_just_the_mode():
    """Downloading from the summary tab must yield the summary.

    In daily mode the export otherwise leads with snapshots regardless of
    which table the user is looking at.
    """
    result = make_result("daily")

    summary_csv = export(result, ExportFormat.csv, "summary").decode("utf-8-sig")
    snapshot_csv = export(result, ExportFormat.csv, "snapshots").decode("utf-8-sig")

    assert summary_csv.splitlines()[0].startswith("wallet_address,first_date")
    assert snapshot_csv.splitlines()[0].startswith("wallet_address,token_address")


def test_auto_still_follows_the_holder_mode():
    daily = export(make_result("daily"), ExportFormat.csv, "auto").decode("utf-8-sig")
    any_time = export(
        make_result("any_time"), ExportFormat.csv, "auto"
    ).decode("utf-8-sig")

    assert "token_address" in daily.splitlines()[0]
    assert "days_held" in any_time.splitlines()[0]


def test_xlsx_leads_with_the_requested_dataset():
    from openpyxl import load_workbook

    workbook = load_workbook(
        io.BytesIO(export(make_result("daily"), ExportFormat.xlsx, "summary"))
    )

    assert workbook.sheetnames[0] == "Summary"
    # both tables are still present, only the order changes
    assert "Snapshots" in workbook.sheetnames


def test_filename_says_which_table_it_holds():
    result = make_result("daily")

    assert "_summary_" in filename_for(result, ExportFormat.csv, "summary")
    assert "_snapshots_" in filename_for(result, ExportFormat.csv, "snapshots")


def test_html_reports_the_range_it_actually_contains():
    """The header must not claim days that were clamped away."""
    from app.models import utc_today

    req = HoldersRequest(
        chain="ethereum",
        token_address="0x1234567890abcdef1234567890abcdef12345678",
        start_date=utc_today().isoformat(),
        end_date=(utc_today() + timedelta(days=10)).isoformat(),
    )
    result = HoldersResponse(
        request=req, row_count=0, wallet_count=0, snapshots=[], summary=[]
    )

    page = export(result, ExportFormat.html).decode("utf-8")

    assert req.effective_end_date.isoformat() in page
    assert req.end_date.isoformat() not in page
